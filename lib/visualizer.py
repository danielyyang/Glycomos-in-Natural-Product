from rdkit import Chem
from rdkit.Chem.Draw import rdMolDraw2D
import logging
from typing import Dict, List, Set, Optional
import pandas as pd
import os
import sys

# 动态添加 lib 路径以支持导入 (Dynamically add lib path to support imports)
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

try:
    import sugar_utils
except ImportError:
    # Fallback if running from root
    from lib import sugar_utils

# 配置日志 (Configure logging)
logger = logging.getLogger(__name__)

class StructureVisualizer:
    """
    负责糖脂结构的解析与可视化 (Handles structure parsing and visualization of glycolipids)
    核心功能：基于 sugar_utils 识别糖环，区分糖基(Glycan)与苷元(Aglycone)，并进行染色可视化。
    """
    
    def __init__(self):
        pass

    def analyze_glycolipid(self, smiles: str, output_path: str = None) -> bool:
        """
        解析糖脂结构并生成可视化图像 (Analyze glycolipid structure and generate visualization)
        
        Args:
            smiles: 分子的 SMILES 字符串
            output_path: 图片保存路径 (可选)
            
        Returns:
            bool: 是否成功生成 (Success status)
        """
        mol = Chem.MolFromSmiles(smiles)
        if not mol:
            logger.error(f"Invalid SMILES: {smiles}")
            return False

        try:
            # 1. 使用 sugar_utils 识别糖单元 (Identify Sugar Units using sugar_utils)
            sugar_units, atom_to_sugar = sugar_utils.get_sugar_units(mol)
            linkages = sugar_utils.find_glycosidic_linkages(mol, sugar_units, atom_to_sugar)
            
            # 2. 收集原子分类 (Collect atom classifications)
            sugar_ring_atoms: Set[int] = set()
            sugar_substituent_atoms: Set[int] = set()
            
            # Helper to check if an atom is part of ANY sugar ring
            all_sugar_ring_atoms = set()
            for unit in sugar_units:
                all_sugar_ring_atoms.update(unit['ring_atoms'])

            for unit in sugar_units:
                # A. Ring Atoms -> Red
                ring_atoms = set(unit['ring_atoms'])
                sugar_ring_atoms.update(ring_atoms)
                
                # B. Substituents -> Yellow
                # Include atoms in position_map (like C6) that are NOT in the ring
                mapped_atoms = set(unit['position_map'].keys())
                exocyclic_carbons = mapped_atoms - ring_atoms
                sugar_substituent_atoms.update(exocyclic_carbons)
                
                # Scan neighbors of ALL mapped atoms (ring + exocyclic) to find functional groups (OH, etc.)
                for atom_idx in mapped_atoms:
                    atom = mol.GetAtomWithIdx(atom_idx)
                    for neighbor in atom.GetNeighbors():
                        n_idx = neighbor.GetIdx()
                        # If neighbor is part of a sugar ring, ignore (it's the ring bond or glycosidic bond to another sugar)
                        if n_idx in all_sugar_ring_atoms:
                            continue
                        
                        # If neighbor is already counted as exocyclic C, ignore
                        if n_idx in sugar_substituent_atoms:
                            continue
                            
                        # If neighbor is NOT in ring, it's likely a substituent (OH, O-Methyl, etc.)
                        # Note: This might capture the Aglycone attachment point (glycosidic O). 
                        # We ideally want the Glycosidic O to be "Substituent" color (Yellow) or special.
                        # Using Yellow for now as it's "attached to sugar ring".
                        sugar_substituent_atoms.add(n_idx)

            # 3. 识别苷元部分 (Identify Aglycone Part)
            all_atom_indices = set(range(mol.GetNumAtoms()))
            # Aglycone = All - (Sugar Rings + Substituents)
            aglycone_atoms = all_atom_indices - sugar_ring_atoms - sugar_substituent_atoms
            
            # 4. 准备着色方案 (Prepare Coloring Scheme)
            # 糖环 (Sugar Ring): 红色 (Red) (1.0, 0.6, 0.6)
            # 糖取代基 (Substituents): 黄色 (Yellow) (1.0, 1.0, 0.6)
            # 苷元 (Aglycone): 蓝色 (Blue) (0.6, 0.8, 1.0)
            
            highlight_atoms = list(all_atom_indices)
            highlight_atom_colors: Dict[int, tuple] = {}
            
            for idx in sugar_ring_atoms:
                highlight_atom_colors[idx] = (1.0, 0.5, 0.5) # Red
            
            for idx in sugar_substituent_atoms:
                highlight_atom_colors[idx] = (1.0, 1.0, 0.4) # Yellow
            
            for idx in aglycone_atoms:
                highlight_atom_colors[idx] = (0.5, 0.8, 1.0) # Blue
                
            # 5. 绘制图像 (Draw Image)
            d2d = rdMolDraw2D.MolDraw2DCairo(600, 400)
            
            # 设置绘图选项 (Drawing Options)
            dopts = d2d.drawOptions()
            # dopts.addAtomIndices = True # DEBUG only
            d2d.DrawMolecule(mol, highlightAtoms=highlight_atoms, highlightAtomColors=highlight_atom_colors)
            d2d.FinishDrawing()
            
            if output_path:
                os.makedirs(os.path.dirname(output_path), exist_ok=True)
                d2d.WriteDrawingText(output_path)
            
            return True
            
        except Exception as e:
            logger.error(f"Error analyzing/drawing molecule: {e}", exc_info=True)
            return False

    def batch_process_from_file(self, file_path: str, output_dir: str = "images/batch/") -> None:
        """
        批量处理文件中的所有 SMILES，并将生成的图片嵌入到 Excel 中 (Batch process all SMILES and embed images into Excel)
        Structure_Image will be placed in the FIRST column (Column A).
        """
        logger.info(f"Starting batch processing from {file_path}...")
        
        if not os.path.exists(file_path):
            logger.error(f"File not found: {file_path}")
            return

        os.makedirs(output_dir, exist_ok=True)
            
        try:
            # 2. 准备 OpenPyXL 进行编辑 (Prepare OpenPyXL for editing)
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.utils import get_column_letter
            
            wb = load_workbook(file_path)
            # Process all sheets
            sheets = wb.sheetnames
            
            for sheet_name in sheets:
                ws = wb[sheet_name]
                
                # Check Header
                header_row_vals = [cell.value for cell in ws[1]]
                
                # Find SMILES column (0-based index in this list)
                try:
                    # Case insensitive search
                    smiles_idx_list = [i for i, h in enumerate(header_row_vals) if h and str(h).lower() in ['smiles', 'canonical_smiles']]
                    if not smiles_idx_list:
                        logger.warning(f"No SMILES column in sheet {sheet_name}")
                        continue
                    original_smiles_col_idx = smiles_idx_list[0] + 1 # 1-based for openpyxl
                except ValueError:
                    continue

                # Insert Image Column at A if not present
                current_img_col_idx = 1
                if header_row_vals[0] != 'Structure_Image':
                    logger.info(f"Inserting Structure_Image column at A in {sheet_name}...")
                    ws.insert_cols(1)
                    ws.cell(row=1, column=1, value='Structure_Image')
                    # Shift SMILES column index
                    original_smiles_col_idx += 1
                
                # Iterate rows
                success_count = 0
                for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    try:
                        smiles_cell = row[original_smiles_col_idx - 1] # tuple index is 0-based
                        smiles = smiles_cell.value
                    except IndexError:
                        continue

                    if not smiles or not isinstance(smiles, str):
                        continue
                    
                    # Generate ID for filename
                    img_name = f"struct_{sheet_name}_{i}"
                    output_image_path = os.path.join(output_dir, f"{img_name}.png")
                    
                    # Generate and Embed
                    if self.analyze_glycolipid(smiles, output_image_path):
                        success_count += 1
                        try:
                            img = XLImage(output_image_path)
                            img.width = 300
                            img.height = 200
                            
                            cell_loc = ws.cell(row=i, column=current_img_col_idx).coordinate
                            ws.add_image(img, cell_loc)
                            ws.row_dimensions[i].height = 150
                        except Exception as e:
                            logger.warning(f"Embedding failed for row {i}: {e}")
                
                # Adjust column width for A
                ws.column_dimensions['A'].width = 40
                logger.info(f"Embedded {success_count} images in {sheet_name}")

            wb.save(file_path)
            logger.info(f"Batch processing complete. Saved to {file_path}")
            
        except Exception as e:
            logger.error(f"Error in batch processing: {e}", exc_info=True)

if __name__ == "__main__":
    # Test with a sample SMILES
    test_smiles = "CO[C@@H]1O[C@H](CO[C@@H]2O[C@H](COC(=O)OCC3c4ccccc4-c4ccccc43)[C@@H](OCc3ccccc3)[C@H](OCc3ccccc3)[C@H]2OC(=O)c2ccccc2)[C@@H](OCc2ccccc2)[C@H](OCc2ccccc2)[C@H]1OC(=O)c1ccccc1"
    viz = StructureVisualizer()
    viz.analyze_glycolipid(test_smiles, "images/test_sugar_utils_viz.png")
    print("Test image generated: images/test_sugar_utils_viz.png")

