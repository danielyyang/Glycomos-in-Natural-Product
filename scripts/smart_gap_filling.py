import pandas as pd
import os
import sys
import requests
import time
import argparse
import urllib.parse
from tqdm import tqdm
from rdkit import Chem
from rdkit.Chem import AllChem
from rdkit import DataStructs
import openpyxl
from openpyxl.styles import Font

def get_morgan_fp(smiles):
    try:
        mol = Chem.MolFromSmiles(smiles)
        if mol:
            return AllChem.GetMorganFingerprintAsBitVect(mol, 2, nBits=2048)
    except:
        pass
    return None

def fill_classification(df):
    """
    Fills 'No Classified' rows based on:
    1. InChIKey (first 14 chars) match with a classified row.
    2. Aglycan_SMILE Morgan Fingerprint Tanimoto similarity >= 0.8 with a classified row.
    Returns (df, imputed_cells_set)
    """
    imputed_cells = set()
    classified_df = df[df['chemical_super_class'] != 'No Classified'].copy()
    unclassified_df = df[df['chemical_super_class'] == 'No Classified'].copy()

    if unclassified_df.empty or classified_df.empty:
        return df, imputed_cells
        
    print(f"Attempting to classify {len(unclassified_df)} unclassified rows...")
    
    # Pre-compute InChIKey blocks and fingerprints for classified rows
    classified_df['inchikey_block1'] = classified_df['standard_inchi_key'].astype(str).str[:14]
    
    # Only compute FP for the first Aglycan part
    def extract_primary_aglycan(s):
        if pd.isna(s) or not str(s).strip(): return ""
        return str(s).split('|')[0].strip()
        
    classified_df['primary_aglycan'] = classified_df['Aglycan_SMILE_ALL'].apply(extract_primary_aglycan)
    tqdm.pandas(desc="Computing Classified FPs")
    classified_fps = classified_df['primary_aglycan'].progress_apply(get_morgan_fp)
    classified_df['fp'] = classified_fps
    valid_classified = classified_df.dropna(subset=['fp']).copy()

    changes = 0
    tqdm.pandas(desc="Inferring Classes")
    
    for idx, row in tqdm(unclassified_df.iterrows(), total=len(unclassified_df), desc="Classifying"):
        # Strategy 1: InChIKey Block 1 Match
        ikey = str(row.get('standard_inchi_key', ''))[:14]
        match = classified_df[classified_df['inchikey_block1'] == ikey]
        if not match.empty:
            inferred_class = match.iloc[0]['chemical_super_class']
            df.at[idx, 'chemical_super_class'] = inferred_class
            imputed_cells.add((idx, 'chemical_super_class'))
            changes += 1
            continue
            
        # Strategy 2: Tanimoto Similarity on Primary Aglycan > 0.8
        primary_ag = extract_primary_aglycan(row.get('Aglycan_SMILE_ALL'))
        if primary_ag:
            fp1 = get_morgan_fp(primary_ag)
            if fp1:
                best_sim = 0
                best_class = None
                for c_idx, c_row in valid_classified.iterrows():
                    fp2 = c_row['fp']
                    sim = DataStructs.TanimotoSimilarity(fp1, fp2)
                    if sim > best_sim:
                        best_sim = sim
                        best_class = c_row['chemical_super_class']
                        
                if best_sim >= 0.8 and best_class:
                    df.at[idx, 'chemical_super_class'] = best_class
                    imputed_cells.add((idx, 'chemical_super_class'))
                    changes += 1

    print(f"Successfully inferred classes for {changes} rows.")
    return df, imputed_cells

# --- Offline Taxonomy Interpolation ---
TAXONOMY_DICT = {
    "Arabidopsis thaliana": "Brassicaceae",
    "Oryza sativa": "Poaceae",
    "Zea mays": "Poaceae",
    "Saccharomyces cerevisiae": "Saccharomycetaceae",
    "Homo sapiens": "Hominidae",
    "Mus musculus": "Muridae",
    "Escherichia coli": "Enterobacteriaceae",
    "Bacillus subtilis": "Bacillaceae",
    "Panax ginseng": "Araliaceae",
    "Glycyrrhiza glabra": "Fabaceae",
    "Astragalus membranaceus": "Fabaceae",
    "Staphylococcus aureus": "Staphylococcaceae",
    "Pseudomonas aeruginosa": "Pseudomonadaceae",
    # Add more known species-to-family mappings here as needed
}

def get_organism_from_pubchem(iupac_name):
    if not iupac_name or pd.isna(iupac_name):
        return None
    try:
        # 1. Translate Name to CID
        url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{urllib.parse.quote(str(iupac_name))}/cids/JSON"
        res = requests.get(url, timeout=10)
        if res.status_code == 200:
            cids = res.json().get('IdentifierList', {}).get('CID', [])
            if cids:
                cid = cids[0]
                # 2. Get Taxonomy from PUG View
                tax_url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{cid}/JSON"
                tax_res = requests.get(tax_url, timeout=10)
                if tax_res.status_code == 200:
                    data = tax_res.json()
                    record = data.get('Record', {})
                    for sec in record.get('Section', []):
                        # Look for 'Associated Disorders and Diseases' or 'Taxonomy' or 'Biomolecular Interactions and Pathways' 
                        # where species/organism might be listed. 'Taxonomy' is preferred.
                        if sec.get('TOCHeading') in ('Taxonomy', 'Associated Disorders and Diseases', 'Biomolecular Interactions and Pathways'):
                            for sub_sec in sec.get('Section', []) + [sec]:
                                for info in sub_sec.get('Information', []):
                                    if 'Name' in info and 'Value' in info:
                                        name = info['Name']
                                        if 'Organism' in name or 'Species' in name or 'Taxonomy' in name:
                                            try:
                                                val = info['Value']['StringWithMarkup'][0]['String']
                                                return val
                                            except:
                                                pass
                                        elif sec.get('TOCHeading') == 'Taxonomy':
                                            try:
                                                val = info['Value']['StringWithMarkup'][0]['String']
                                                return val
                                            except:
                                                pass
        time.sleep(0.3)
    except Exception as e:
        # Silent fail for individual api misses
        pass
    return None

def get_family_from_gbif(organism):
    if not organism or pd.isna(organism):
        return None
    try:
        url = f"https://api.gbif.org/v1/species/match?name={urllib.parse.quote(str(organism))}"
        res = requests.get(url, timeout=10)
        if res.status_code == 200:
            data = res.json()
            return data.get('family', None)
        time.sleep(0.1)
    except Exception as e:
        pass
    return None

def fill_taxonomy_online(df):
    """
    Fills 'organisms' and 'Family' using online databases.
    - organisms: Fetched from PubChem Taxonomy via iupac_name
    - Family: Fetched from GBIF Species match via organisms
    Flags rows that were imputed so they can be colored red in Excel.
    """
    if 'Family' not in df.columns:
        if 'organisms' in df.columns:
            org_idx = df.columns.get_loc('organisms')
            df.insert(org_idx + 1, 'Family', "")
        else:
            df['organisms'] = ""
            df['Family'] = ""
            
    imputed_cells = set()
    changes_org = 0
    changes_fam = 0
    
    print("Applying Online Taxonomy Resolution (PubChem & GBIF)...")
    
    pubchem_cache = {}
    gbif_cache = {}
    
    for idx, row in tqdm(df.iterrows(), total=len(df), desc="Fetching Taxonomy from API"):
        org = str(row.get('organisms', '')).strip()
        fam = str(row.get('Family', '')).strip()
        iupac = str(row.get('iupac_name', '')).strip()
        
        # 1. Fill Organism
        if (not org or org == 'nan') and iupac and iupac != 'nan':
            if iupac in pubchem_cache:
                inferred_org = pubchem_cache[iupac]
            else:
                inferred_org = get_organism_from_pubchem(iupac)
                pubchem_cache[iupac] = inferred_org
                
            if inferred_org:
                df.at[idx, 'organisms'] = inferred_org
                imputed_cells.add((idx, 'organisms'))
                org = inferred_org # update org for next step
                changes_org += 1
                
        # 2. Fill Family
        if org and org != 'nan' and (not fam or fam == 'nan'):
            if org in TAXONOMY_DICT:
                df.at[idx, 'Family'] = TAXONOMY_DICT[org]
                imputed_cells.add((idx, 'Family'))
                changes_fam += 1
            else:
                if org in gbif_cache:
                    inferred_fam = gbif_cache[org]
                else:
                    inferred_fam = get_family_from_gbif(org)
                    gbif_cache[org] = inferred_fam
                    
                if inferred_fam:
                    df.at[idx, 'Family'] = inferred_fam
                    imputed_cells.add((idx, 'Family'))
                    changes_fam += 1
                        
    print(f"Successfully inferred {changes_org} organisms and {changes_fam} Family taxonomy rows via API.")
    return df, imputed_cells

def main():
    parser = argparse.ArgumentParser(description="Glycan Database: Smart Classification & Gap Filling")
    # limit arg kept for backward compatibility if needed
    parser.add_argument("--limit", type=int, default=0, help="Ignored. API gap filling is disabled.")
    args = parser.parse_args()

    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    input_csv = os.path.join(base_dir, "reports", "Coconut_Sugar_Final.csv")
    output_csv = os.path.join(base_dir, "reports", "Coconut_Sugar_Enriched.csv")
    output_excel = os.path.join(base_dir, "reports", "Coconut_Sugar_Enriched.xlsx")

    if not os.path.exists(input_csv):
        print(f"Error: {input_csv} not found. Run unified_pipeline.py first.")
        return

    print(f"Loading dataset from {input_csv}...")
    df = pd.read_csv(input_csv, low_memory=False)

    total_imputed = set()

    print("\n--- Phase 3.1: Smart Classification ---")
    df, class_imputed = fill_classification(df)
    total_imputed.update(class_imputed)

    print("\n--- Phase 3.2: Offline Taxonomy Resolution ---")
    df, tax_imputed = fill_taxonomy_offline(df)
    total_imputed.update(tax_imputed)

    print(f"\nSaving enriched dataset CSV to {output_csv}...")
    df.to_csv(output_csv, index=False)
    
    # ---------------------------------------------------------
    # Generate Stylized Excel Output
    # ---------------------------------------------------------
    print(f"Generating Styled Excel Output to {output_excel}...")
    # First write basic to excel without index
    df.to_excel(output_excel, index=False)
    
    # Now load with openpyxl to apply styles
    wb = openpyxl.load_workbook(output_excel)
    ws = wb.active
    
    # Build a lookup for column names to column numbers (1-indexed in excel)
    col_name_to_idx = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    
    red_font = Font(color="FF0000")
    
    # total_imputed holds (dataframe_index, column_name)
    # DataFrame index + 2 = Excel row (1 for header, 1 for 0-indexing)
    painted_cells = 0
    for df_idx, col_name in total_imputed:
        if col_name in col_name_to_idx:
            excel_row = df_idx + 2
            excel_col = col_name_to_idx[col_name]
            ws.cell(row=excel_row, column=excel_col).font = red_font
            painted_cells += 1
            
    wb.save(output_excel)
    print(f"Excel styling complete. Painted {painted_cells} cells red.")
    print("Pipeline Complete Phase 3!")

if __name__ == "__main__":
    main()
